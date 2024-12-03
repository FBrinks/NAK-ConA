import re
import os
import subprocess
import platform
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QWidget, QVBoxLayout, QLabel, QPushButton, QLineEdit

# Color mapping
color_mapping = {
    "yellow": "FFFF00",
    "red": "FF0000",
    "green": "00FF00",
    "grey": "808080",
    "purple": "800080"
}

def open_file(file_path):
    if platform.system() == "Windows":
        os.startfile(file_path)
    elif platform.system() == "Darwin":  # macOS
        subprocess.call(["open", file_path])
    else:
        print("Error: Unsupported OS. Cannot open file.")

def load_and_process_excel(file_path):
    df = pd.read_excel(file_path)
    print("Excel file loaded successfully.")
    return df

def map_columns(df, columns_mapping): # Map the columns in the DataFrame to the standard column names
    """ Map the columns in the DataFrame to the standard column names """
    selected_columns = {}
    for standard_name, possible_names in columns_mapping.items():
        for name in possible_names:
            if name in df.columns:
                selected_columns[standard_name] = name
                break
    print(f"Found columns: {selected_columns}")
    return selected_columns

def clean_external_color_code(df): # Remove special characters from 'External color code'
    df["External color code"] = df["External color code"].astype(str).apply(lambda x: re.sub(r"[*-]", "", x))
    print("Cleaned 'External color code' values:")
    print(df["External color code"].head().to_string(index=False))

def create_imagebank_search_column(row, brand_name, supplier_no_handler, combiner):
    if row["Brand"] == brand_name:
        modified_supplier_no = supplier_no_handler(row["Supplier product no"])
        return combiner(modified_supplier_no, row["External color code"])
    return None

def process_imagebank_search(df, ws): # Apply Brand-specific logic to create 'Imagebank search' column
    brand_handlers = {
       "The North Face": {
              "supplier_no_handler": lambda x: x[4:], # Remove first 4 characters
              "combiner": lambda supplier_no, color_code: supplier_no + color_code # Combine supplier number and color code
       },
       "Marmot": {
                "supplier_no_handler": lambda x: x[1:] if x and x[0].lower() == 'm' else x, # Remove first character
                "combiner": lambda supplier_no, color_code: f'"{supplier_no}-{color_code}"' # Combine supplier number and color code
       },
    }
   
    combined_search_values = []
    for brand_name, handlers in brand_handlers.items():
       if df["Brand"].str.contains(brand_name).any():
           print(f"'{brand_name}' found in 'Brand' column. Processing 'Imagebank search' column.")
           df["Imagebank search"] = df.apply(
               create_imagebank_search_column,
                axis=1,
                brand_name=brand_name,
                supplier_no_handler=handlers["supplier_no_handler"],
                combiner=handlers["combiner"]
        )
           print(f"'Imagebank search' column processed for '{brand_name}'.")
           combined_search_values.extend(df["Imagebank search"].dropna().tolist())
           
    if combined_search_values:
        combined_search_values_str = ' '.join(combined_search_values)
        new_column_index = ws.max_column + 1
        ws.cell(row=1, column=new_column_index).value = "Imagebank search"
        ws.cell(row=2, column=new_column_index).value = combined_search_values_str
        ws.cell(row=1, column=new_column_index).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=new_column_index).alignment = Alignment(horizontal="center", vertical="center")
        print(f"'Imagebank search' column added to the Excel file.")
    else:
        print(f"No brands found in 'Brand' column. Skipping 'Imagebank search' column processing.")
               
           
def apply_color_fill(ws, color_mapping, s1_s4_columns): # Apply color fill to 'S1', 'S2', 'S3', 'S4' columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if ws.cell(row=1, column=cell.column).value in s1_s4_columns:
                cell_value = str(cell.value).strip().lower()
                if cell_value in color_mapping:
                    cell.fill = PatternFill(start_color=color_mapping[cell_value], end_color=color_mapping[cell_value], fill_type="solid")

def format_product_item_number(ws): # Format 'Product/item number' column as text
    for cell in ws['A'][1:]:
        cell.number_format = '0'

def adjust_column_widths(ws): # Adjust column widths based on the longest value in each column
    for col in ws.columns:
        values = [str(cell.value) for cell in col if cell.value]
        if values:
            max_length = max(len(value) for value in values) + 2
            ws.column_dimensions[col[0].column_letter].width = max_length

def align_cells(ws): # Align cells to center
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

def process_excel(file_path, save_path):
    try:
        df = load_and_process_excel(file_path)
        columns_mapping = {
            "Product/item number": ["Product/item number", "Artikelnummer"],
            "Stock OneStock total": ["Stock OneStock total", "Lager OneStock sum"],
            "S1": ["S1"],
            "S2": ["S2"],
            "S3": ["S3"],
            "S4": ["S4"],
            "Brand": ["Brand", "Varumärke"],
            "Product name": ["Product name", "Produktnamn"],
            "Product": ["Product", "Produkt"],
            "Supplier product no": ["Supplier product no", "Leverantörens produktnr."],
            "External color code": ["External color code", "Extern färgkod"],
        }
        selected_columns = map_columns(df, columns_mapping)
        new_df = df[selected_columns.values()].rename(columns={v: k for k, v in selected_columns.items()})
        print("Columns selected and renamed.")
        new_df.drop_duplicates(subset=["Product", "Supplier product no", "External color code"], inplace=True) # Remove duplicates
        print("Duplicates removed.")
        clean_external_color_code(new_df) # Clean 'External color code' values
        new_df.to_excel(save_path, index=False)
        print(f"New Excel file created and saved at {save_path}.")
        wb = load_workbook(save_path)
        ws = wb.active
        process_imagebank_search(new_df, ws)
        apply_color_fill(ws, color_mapping, ["S1", "S2", "S3", "S4"])
        format_product_item_number(ws)
        adjust_column_widths(ws)
        align_cells(ws)
        wb.save(save_path)
        print(f"Styled Excel file saved at {save_path}.")
        return True
    except Exception as e:
        print(f"ERROR! Try again with an exported items-list {e}")
        return False

class ExcelProcessingWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.apply_styles()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)

        # Title
        title = QLabel("Excel File Processor for Image Bank Search")
        title.setObjectName("title")
        layout.addWidget(title)

        # Info section
        info_text = """
        How to use this feature:
        
        1. Click 'Choose File' to select an Excel file exported from OneStock
        2. The file should contain the following columns:
           - Product/item number
           - Stock OneStock total
           - S1, S2, S3, S4
           - Brand
           - Product name
           - Supplier product no
           - External color code
           
        3. Click 'Process' to generate a formatted Excel file
        4. The processed file will be saved with '_ConA' suffix
        
        The tool will:
        • Remove duplicates
        • Clean color codes
        • Create image bank search strings
        • Apply color coding to S1-S4 columns
        • Format numbers and align cells
        """
        info_label = QLabel(info_text)
        info_label.setObjectName("info")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # File selection section
        file_section = QWidget()
        file_layout = QVBoxLayout(file_section)
        file_layout.setSpacing(5)

        file_label = QLabel("Selected Excel file:")
        file_label.setObjectName("section-label")
        file_layout.addWidget(file_label)

        self.file_path_display = QLineEdit(self)
        self.file_path_display.setReadOnly(True)
        self.file_path_display.setPlaceholderText("No file selected")
        file_layout.addWidget(self.file_path_display)

        # Buttons
        button_layout = QVBoxLayout()
        button_layout.setSpacing(10)

        self.choose_file_button = QPushButton("Choose Excel File", self)
        self.choose_file_button.setObjectName("choose-button")
        self.choose_file_button.clicked.connect(self.choose_file)
        button_layout.addWidget(self.choose_file_button)

        self.process_button = QPushButton("Process File", self)
        self.process_button.setObjectName("process-button")
        self.process_button.clicked.connect(self.process_file)
        button_layout.addWidget(self.process_button)

        file_layout.addLayout(button_layout)
        layout.addWidget(file_section)

        self.setLayout(layout)

    def apply_styles(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f5f5;
                font-family: Arial;
            }
            #title {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px;
            }
            #info {
                background-color: #ffffff;
                padding: 15px;
                border-radius: 5px;
                border: 1px solid #dcdcdc;
                color: #34495e;
                margin: 10px 0;
            }
            #section-label {
                font-weight: bold;
                color: #2c3e50;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
            QPushButton {
                padding: 8px 15px;
                border-radius: 4px;
                border: none;
                color: white;
                font-weight: bold;
            }
            #choose-button {
                background-color: #3498db;
            }
            #choose-button:hover {
                background-color: #2980b9;
            }
            #process-button {
                background-color: #2ecc71;
            }
            #process-button:hover {
                background-color: #27ae60;
            }
        """)

    def choose_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel files (*.xlsx *.xls)")
        if file_path:
            self.file_path_display.setText(file_path)

    def process_file(self):
        file_path = self.file_path_display.text()
        if not file_path:
            QMessageBox.warning(self, "Warning", "No file chosen.")
            return

        save_path = os.path.splitext(file_path)[0] + "_ConA.xlsx"
        success = process_excel(file_path, save_path)
        if success:
            QMessageBox.information(self, "Success", f"File processed and saved successfully as {save_path}.")
        else:
            QMessageBox.warning(self, "Error", "Failed to process the file.")

